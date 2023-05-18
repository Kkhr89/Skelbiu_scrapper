from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import Select
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font
import matplotlib.pyplot as plt


# Initialize variables:
url = 'https://www.skelbiu.lt/skelbimai/nekilnojamasis-turtas/nuoma/ilgalaike/butu-nuoma/'
flat_dict = dict()
prices_set, square_set = [], []

# Request input from terminal about flats search parameters:
min_price = int(input('Input MIN price:'))
max_price = int(input('Input MAX price:'))
min_space = int(input('Input MIN required space in m^2:'))
max_space = int(input('Input MAX required space in m^2:'))
min_rooms = input('Input MIN required rooms amount:')
max_rooms = input('Input MAX required rooms amount:')

# Driver Manager initialization & Chromedriver launch:
service = Service(executable_path=ChromeDriverManager().install())
options = webdriver.ChromeOptions()
options.add_argument("--disable-dev-shm-usage")
options.add_argument("--no-sandbox")
browser = webdriver.Chrome(service=service, options=options)
browser.maximize_window()
wait = WebDriverWait(browser, 10)

# Go through page:
browser.get(url)
wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'button#onetrust-accept-btn-handler'))).click()
wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, '.fa-caret-down'))).click()
wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, '#cityFor119'))).click()
wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'input#addCitiesButton'))).click()
wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, '#moreFilters'))).click()
wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'input[name="cost_min"]')))\
    .send_keys(min_price)
wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'input[name="cost_max"]')))\
    .send_keys(max_price)
wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'input[name="space_min"]')))\
    .send_keys(min_space)
wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'input[name="space_max"]')))\
    .send_keys(max_space)
Select(wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'select[name="rooms_min"]'))))\
    .select_by_value(min_rooms)
Select(wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'select[name="rooms_max"]'))))\
    .select_by_value(max_rooms)
wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, '.filter>button'))).click()

# Initialize excel file and make initial template:
wb = Workbook()
ws1 = wb.active
ws1.title = 'Summary'
ws1['A1'] = 'Id'
ws1['B1'] = 'Price, eur'
ws1['C1'] = 'Square, m^2'
ws1['D1'] = 'Rooms amount'
ws1['E1'] = 'Url'
ws1.column_dimensions['A'].width = 12
ws1.column_dimensions['B'].width = 8
ws1.column_dimensions['C'].width = 11
ws1.column_dimensions['D'].width = 13
ws1.column_dimensions['E'].width = 100
ws2 = wb.create_sheet('Statistics')
ws2['A1'] = 'Average price, eur:'
ws2['A2'] = 'Max price:'
ws2['A3'] = 'Min price:'
ws2['A4'] = 'Total flats amount at skelbiu:'
ws2['A5'] = 'Average square, m^2:'
ws2.column_dimensions['A'].width = 25
ws2.column_dimensions['B'].width = 25
ws3 = wb.create_sheet('Chart')
for cell in ws1[1]:
    cell.font = Font(bold=True)

while True:
    try:
        next_btn = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'a[rel="next"]')))
        # Parse resulted html:
        html_content = browser.page_source
        soup = BeautifulSoup(html_content, 'html.parser')

        # Prepare dict with all flats rough data and sets for further calculations from html:
        flats = soup.find_all(class_='simpleAds')
        for flat in flats:
            price = int(flat.select(".adsPrice>span:first-child")[0].get_text().split()[0])
            square = float(flat.select(".adsTextMoreDetails")[0].get_text().split()[0])
            rooms = int(flat.select(".adsTextMoreDetails")[0].get_text().split()[3])
            urls = flat.select('a.js-cfuser-link')[0].get("href")
            flat_dict[flat.get('id')] = {'Price': price,
                                         'Square': square,
                                         'Rooms': rooms,
                                         'URL': f'{urls}'
                                         }
            prices_set.append(price)
            square_set.append(square)

        # Move to the next page:
        next_btn.click()
    except :
        break

browser.close()
browser.quit()

# Write from flat_dict to excel file:
i = 2
for key, value in flat_dict.items():
    ws1[f'A{i}'] = key
    ws1[f'B{i}'] = value['Price']
    ws1[f'C{i}'] = value['Square']
    ws1[f'D{i}'] = value['Rooms']
    ws1[f'E{i}'] = f"https://www.skelbiu.lt{value['URL']}"
    i += 1

# Calculate Statistics and write to excel file:
ws2['B1'] = sum(prices_set) / len(prices_set)
ws2['B2'] = max(prices_set)
ws2['B3'] = min(prices_set)
ws2['B4'] = len(prices_set)
ws2['B5'] = sum(square_set) / len(square_set)

ws1.auto_filter.ref = ws1.dimensions

# Create & add Price Chart to excel file:
plt.hist(prices_set)
plt.title('Prices range')
plt.xlabel('Price, eur')
plt.ylabel('Flats amount')
plt.savefig('price_chart.png')
img = Image('price_chart.png')
ws3.add_image(img, 'A1')

wb.save('skelbiu_scraped.xlsx')
